import urllib.request
#import numpy
import requests
import pandas as pd
import xlwt
import datetime
from bs4 import BeautifulSoup
from xlwt import Workbook
import openpyxl

# `````1 Code for kotak-standard-multicap-fund-direct-plan 
url1='https://www.moneycontrol.com/mutual-funds/nav/kotak-standard-multicap-fund-direct-plan/MKM520'
response1=requests.get(url1)

# To see if we are able to read the website without error. <Response [200]> means we are able to access the web page.
print(response1)

# Now we will parse the reponse text that we got html format using BeautifulSoup
soup1=BeautifulSoup(response1.text,"html.parser")
#print(soup1)

# We will use findAll function to find a aprticular tag for all appearances.
l1=soup1.findAll('span',{"class":"amt"})[0].text
l2=soup1.findAll('div',{"class":"grayvalue"})[0].text
Date1=''.join(map(str, l2))
NAV1=''.join(map(str, l1))
print(NAV1[1:])
print(Date1[6:-1])
df1=pd.read_excel('Myportfolio.xlsx')
# define workbook to write the Daily NAV
wb=openpyxl.load_workbook(filename='Myportfolio.xlsx')
mf1 = wb.worksheets[0]
max_row1 = mf1.max_row
#print(max_row1)

for i in range(max_row1):
 if( str(Date1[6:-1]) not in str(df1['Date'])):
  mf1.cell(row=max_row1+1, column=1).value=Date1[6:-1]
  mf1.cell(row=max_row1+1, column=2).value=NAV1[1:]
wb.save("Myportfolio.xlsx")

#````2 Code for sbi-blue-chip-fund

url2='https://www.moneycontrol.com/mutual-funds/nav/sbi-blue-chip-fund/MSB079'
response2=requests.get(url2)

# To see if we are able to read the website without error. <Response [200]> means we are able to access the web page.
print(response2)

# Now we will parse the reponse text that we got html format using BeautifulSoup
soup2=BeautifulSoup(response2.text,"html.parser")
#print(soup2)

# We will use findAll function to find a aprticular tag for all appearances.
l3=soup2.findAll('span',{"class":"amt"})[0].text
l4=soup2.findAll('div',{"class":"grayvalue"})[0].text
Date2=''.join(map(str, l4))
NAV2=''.join(map(str, l3))
print(NAV2[1:])
print(Date2[6:-1])

mf2 = wb.worksheets[1]
max_row2 = mf2.max_row
#print(max_row2)

for i in range(max_row2):
 if( str(Date2[6:-1]) not in str(df1['Date'])):
   mf2.cell(row=max_row2+1, column=1).value=Date2[6:-1]
   mf2.cell(row=max_row2+1, column=2).value=NAV2[1:]
wb.save("Myportfolio.xlsx")

#``````3 Code for reliance-small-cap-fund

url3='https://www.moneycontrol.com/mutual-funds/nav/reliance-small-cap-fund/MRC587'
response3=requests.get(url3)

# To see if we are able to read the website without error. <Response [200]> means we are able to access the web page.
print(response3)

# Now we will parse the reponse text that we got html format using BeautifulSoup
soup3=BeautifulSoup(response3.text,"html.parser")
#print(soup3)

# We will use findAll function to find a aprticular tag for all appearances.
l5=soup3.findAll('span',{"class":"amt"})[0].text
l6=soup3.findAll('div',{"class":"grayvalue"})[0].text
Date3=''.join(map(str, l6))
NAV3=''.join(map(str, l5))
print(NAV3[1:])
print(Date3[6:-1])

mf3 = wb.worksheets[2]
max_row3 = mf3.max_row
#print(df1['Date'])

for i in range(max_row3):
 if( str(Date3[6:-1]) not in str(df1['Date'])):
   mf3.cell(row=max_row3+1, column=1).value=Date3[6:-1]
   mf3.cell(row=max_row3+1, column=2).value=NAV3[1:]
wb.save("Myportfolio.xlsx")

#`````4 Code for aditya-birla-sun-life-small-and-midcap-fund-direct-plan

url4='https://www.moneycontrol.com/mutual-funds/nav/aditya-birla-sun-life-small-and-midcap-fund-direct-plan/MBS833'
response4=requests.get(url4)

# To see if we are able to read the website without error. <Response [200]> means we are able to access the web page.
print(response4)

# Now we will parse the reponse text that we got html format using BeautifulSoup
soup4=BeautifulSoup(response4.text,"html.parser")
#print(soup4)

# We will use findAll function to find a aprticular tag for all appearances.
l7=soup4.findAll('span',{"class":"amt"})[0].text
l8=soup4.findAll('div',{"class":"grayvalue"})[0].text
Date4=''.join(map(str, l8))
NAV4=''.join(map(str, l7))
print(NAV4[1:])
print(Date4[6:-1])

mf4 = wb.worksheets[3]
max_row4 = mf4.max_row
#print(max_row4)

for i in range(max_row4):
 if( str(Date4[6:-1]) not in str(df1['Date'])):
   mf4.cell(row=max_row4+1, column=1).value=Date4[6:-1]
   mf4.cell(row=max_row4+1, column=2).value=NAV4[1:]
wb.save("Myportfolio.xlsx")

#``````5 Code for aditya-birla-sun-life-tax-relief-96

url5='https://www.moneycontrol.com/mutual-funds/nav/aditya-birla-sun-life-tax-relief-96/MBS291'
response5=requests.get(url5)

# To see if we are able to read the website without error. <Response [200]> means we are able to access the web page.
print(response5)

# Now we will parse the reponse text that we got html format using BeautifulSoup
soup5=BeautifulSoup(response5.text,"html.parser")
#print(soup5)

# We will use findAll function to find a aprticular tag for all appearances.
l9=soup5.findAll('span',{"class":"amt"})[0].text
l10=soup5.findAll('div',{"class":"grayvalue"})[0].text
Date5=''.join(map(str, l10))
NAV5=''.join(map(str, l9))
print(NAV5[1:])
print(Date5[6:-1])

mf5 = wb.worksheets[4]
max_row5 = mf5.max_row
#print(max_row5)

for i in range(max_row5):
 if( str(Date5[6:-1]) not in str(df1['Date'])):
   mf5.cell(row=max_row5+1, column=1).value=Date5[6:-1]
   mf5.cell(row=max_row5+1, column=2).value=NAV5[1:]
wb.save("Myportfolio.xlsx")